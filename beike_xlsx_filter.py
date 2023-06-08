import openpyxl


class ErshoufangXlsxFilter(object):
    def __init__(self):
        pass

    def write_cow_to_new_xlsx(self, sheet_new, source_sheet, sheet_num, cow):
        for column in "ABCDEFGHI":
            sheet_new[str(column) + str(sheet_num)].value = source_sheet[str(column) + str(cow)].value

    def xlsx_filter_with_area(self):
        wb = openpyxl.load_workbook("remove_data_1-200.xlsx")
        sheet = wb.active
        area_value_list = []
        for cow in range(2, 2419):
            area_value = sheet["F" + str(cow)].value
            if area_value not in area_value_list:
                area_value_list.append(area_value)

        new_wb = openpyxl.Workbook()
        sheet_num = 0
        for sheet_name in area_value_list:
            new_wb.create_sheet(sheet_name, sheet_num)
            sheet_num += 1

        sheet_0 = new_wb["高新北区"]
        sheet_1 = new_wb["经开北区"]
        sheet_2 = new_wb["南关区"]
        sheet_3 = new_wb["绿园区"]
        sheet_4 = new_wb["二道区"]
        sheet_5 = new_wb["高新区"]
        sheet_6 = new_wb["经开区"]
        sheet_7 = new_wb["宽城区"]
        sheet_8 = new_wb["朝阳区"]
        sheet_9 = new_wb["净月区"]
        sheet_10 = new_wb["汽车产业开发区"]

        sheet_0_num = 2
        sheet_1_num = 2
        sheet_2_num = 2
        sheet_3_num = 2
        sheet_4_num = 2
        sheet_5_num = 2
        sheet_6_num = 2
        sheet_7_num = 2
        sheet_8_num = 2
        sheet_9_num = 2
        sheet_10_num = 2

        sheet_list = [sheet_0, sheet_1, sheet_2, sheet_3, sheet_4, sheet_5, sheet_6, sheet_7, sheet_8, sheet_9, sheet_10]
        sheet_1cow_name_list = ["标题", "总价", "元/平米", "房屋面积", "小区名称", "所在区域", "房屋类型", "房屋楼层", "抵押信息"]

        for main_sheet in sheet_list:
            name_list_num = -1
            for column in "ABCDEFGHI":
                name_list_num += 1
                main_sheet[str(column) + "1"].value = sheet_1cow_name_list[name_list_num]
        for cow in range(2, 2419):
            if sheet["F" + str(cow)].value == "高新北区":
                self.write_cow_to_new_xlsx(sheet_0, sheet, sheet_0_num, cow)
                sheet_0_num += 1
            elif sheet["F" + str(cow)].value == "经开北区":
                self.write_cow_to_new_xlsx(sheet_1, sheet, sheet_1_num, cow)
                sheet_1_num += 1
            elif sheet["F" + str(cow)].value == "南关区":
                self.write_cow_to_new_xlsx(sheet_2, sheet, sheet_2_num, cow)
                sheet_2_num += 1
            elif sheet["F" + str(cow)].value == "绿园区":
                self.write_cow_to_new_xlsx(sheet_3, sheet, sheet_3_num, cow)
                sheet_3_num += 1
            elif sheet["F" + str(cow)].value == "二道区":
                self.write_cow_to_new_xlsx(sheet_4, sheet, sheet_4_num, cow)
                sheet_4_num += 1
            elif sheet["F" + str(cow)].value == "高新区":
                self.write_cow_to_new_xlsx(sheet_5, sheet, sheet_5_num, cow)
                sheet_5_num += 1
            elif sheet["F" + str(cow)].value == "经开区":
                self.write_cow_to_new_xlsx(sheet_6, sheet, sheet_6_num, cow)
                sheet_6_num += 1
            elif sheet["F" + str(cow)].value == "宽城区":
                self.write_cow_to_new_xlsx(sheet_7, sheet, sheet_7_num, cow)
                sheet_7_num += 1
            elif sheet["F" + str(cow)].value == "朝阳区":
                self.write_cow_to_new_xlsx(sheet_8, sheet, sheet_8_num, cow)
                sheet_8_num += 1
            elif sheet["F" + str(cow)].value == "净月区":
                self.write_cow_to_new_xlsx(sheet_9, sheet, sheet_9_num, cow)
                sheet_9_num += 1
            elif sheet["F" + str(cow)].value == "汽车产业开发区":
                self.write_cow_to_new_xlsx(sheet_10, sheet, sheet_10_num, cow)
                sheet_10_num += 1

        new_wb.save("分区域归类Sheet.xlsx")

    def remove_repeat_data(self):
        remove_data_wb = openpyxl.load_workbook("金水区_贝壳网二手房数据集_1-100页.xlsx")
        remove_data_ws = remove_data_wb.active
        ws_name_list = []
        creat_wb = openpyxl.Workbook()
        creat_ws = creat_wb.active
        creat_ws_num = 2
        for remove_data_cow in range(2, 3002):
            if remove_data_ws["A" + str(remove_data_cow)].value not in ws_name_list:
                ws_name_list.append(remove_data_ws["A" + str(remove_data_cow)].value)
                for column in "ABCDEFGHIJK":
                    creat_ws[str(column) + str(creat_ws_num)].value = remove_data_ws[str(column) + str(remove_data_cow)].value
                creat_ws_num += 1
            elif remove_data_ws["A" + str(remove_data_cow)].value in ws_name_list:
                continue

        creat_wb.save("remove_data_1-200.xlsx")


if __name__ == '__main__':
    ErshoufangXlsxFilter().xlsx_filter_with_area()
